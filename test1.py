# -*- coding: utf-8 -*-
"""
ë¯¸ì…ê³  KPI ëŒ€ì‹œë³´ë“œ (Streamlit Â· Robust Reader Â· ì—…ë¡œë“œ íˆìŠ¤í† ë¦¬ 'ì˜êµ¬ ì €ì¥' + ì‚­ì œ ë²„íŠ¼)
- ë”ì¡´ ë°œì£¼í˜„í™© ì—…ë¡œë“œ â†’ KPI/ì°¨íŠ¸/í•„í„° â†’ ì—‘ì…€ ë³´ê³ ì„œ(ìš”ì•½+ìƒì„¸+ì›ë³¸) ë‹¤ìš´ë¡œë“œ
- ì—…ë¡œë“œ íˆìŠ¤í† ë¦¬: ./uploads í´ë”ì— íŒŒì¼ì„ ì €ì¥í•´ 'ì¬ì ‘ì†í•´ë„' ëª©ë¡ì—ì„œ ì„ íƒ/ì‚­ì œ ê°€ëŠ¥
- ë‹¤ì¤‘ ì—…ë¡œë“œ ê°€ëŠ¥(ë™ì¼ íŒŒì¼ëª…ì€ ìµœì‹ ë³¸ìœ¼ë¡œ ì¹˜í™˜), ì„ íƒ ì‚­ì œ ë²„íŠ¼ ì œê³µ
- ë³¸ë¬¸ ë¡œì§ì€ ê¸°ì¡´ test2.pyë¥¼ í† ëŒ€ë¡œ ë³´ê°•(íŒŒì¼ ì €ì¥/ë¶ˆëŸ¬ì˜¤ê¸°ë§Œ ì¶”ê°€)
"""

import io, os, csv, calendar, time, json, re
from pathlib import Path
from datetime import date, datetime
import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st


def halt_app():
    """Safely stop execution both in Streamlit runtime and bare execution."""
    try:
        st.stop()
    except Exception:
        raise SystemExit(0)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ ê¸°ë³¸ ì„¤ì • â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.set_page_config(page_title="ë¯¸ì…ê³  KPI ëŒ€ì‹œë³´ë“œ(Pro)", page_icon="ğŸ“¦", layout="wide")

ROOT_DIR   = Path(__file__).parent if "__file__" in globals() else Path(".")
UPLOAD_DIR = ROOT_DIR / "uploads"
MANIFEST   = UPLOAD_DIR / "manifest.json"
UPLOAD_DIR.mkdir(exist_ok=True)

def month_end(d: date) -> date:
    return date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])

def add_months(d: date, n: int) -> date:
    y = d.year + (d.month - 1 + n) // 12
    m = (d.month - 1 + n) % 12 + 1
    return date(y, m, min(d.day, calendar.monthrange(y, m)[1]))

# ë”ì¡´ ì»¬ëŸ¼ í›„ë³´ & ìƒíƒœ ë¼ë²¨(ì‚¬ì´ë“œë°”ì—ì„œ ë°”ê¿€ ìˆ˜ ìˆìŒ)
COL = {
    "group":   ["ì œí’ˆêµ°(1)","ì œí’ˆêµ°","í’ˆëª©êµ°","ëŒ€ë¶„ë¥˜"],
    "po_no":   ["ë°œì£¼ë²ˆí˜¸","POë²ˆí˜¸","PO_NO","ì£¼ë¬¸ë²ˆí˜¸"],
    "po_date": ["ë°œì£¼ì¼ì","ë°œì£¼ì¼","PO_DATE","ì£¼ë¬¸ì¼ì"],
    "item":    ["í’ˆëª©ëª…","í’ˆëª©","ë‚´ì—­","ITEM_NAME"],
    "vendor":  ["ê±°ë˜ì²˜ëª…","ê±°ë˜ì²˜","ê³µê¸‰ì‚¬"],
    "pgroup":  ["êµ¬ë§¤ê·¸ë£¹","êµ¬ë§¤ê·¸ë£¹ëª…","êµ¬ë§¤ ê·¸ë£¹","Buyer Group","êµ¬ë§¤ê·¸ë£¹ì½”ë“œ"],
    "due":     ["ë°œì£¼ë‚©ê¸°ì¼ì","ë‚©ê¸°ì¼ì","ë‚©ê¸°ì¼","DUE_DATE"],
    "rcv_date":["ì…ê³ ì¼ì","ì…ê³ ì¼","RCV_DATE"],
    "po_qty":  ["ë°œì£¼ìˆ˜ëŸ‰","POìˆ˜ëŸ‰","ë°œì£¼ ìˆ˜ëŸ‰"],
    "rcv_qty": ["ì…ê³ ìˆ˜ëŸ‰","RCVìˆ˜ëŸ‰","ê²€ì‚¬í•©ê²©ìˆ˜ëŸ‰","ì‹¤ì…ê³ ìˆ˜ëŸ‰"],
    "status":  ["ì…ê³ êµ¬ë¶„","ì…ê³ ìƒíƒœ","ì…ê³ ìƒíƒœëª…","ì§„í–‰ìƒíƒœ"],
}
COMPLETE = {"ì…ê³ ì™„ë£Œ","ì™„ë£Œ"}
PARTIAL  = {"ë¶€ë¶„ì…ê³ ","ë¶€ë¶„","ë¶€ë¶„ì™„ë£Œ"}
OPEN     = {"ë¯¸ì…ê³ ","ëŒ€ê¸°","ë¯¸ì™„ë£Œ"}
OPEN_OR_PARTIAL = OPEN | PARTIAL

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ ì—…ë¡œë“œ íˆìŠ¤í† ë¦¬ (ì˜êµ¬ ì €ì¥) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _slug(s: str) -> str:
    s = re.sub(r"[^\w.\-ê°€-í£ ]+", "_", s).strip()
    s = re.sub(r"\s+", "_", s)
    return s[:140] if len(s) > 140 else s

def _load_manifest() -> list:
    if MANIFEST.exists():
        try:
            return json.loads(MANIFEST.read_text("utf-8"))
        except Exception:
            return []
    return []

def _save_manifest(rows: list):
    MANIFEST.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

def list_uploads() -> list:
    rows = _load_manifest()
    rows = [r for r in rows if (UPLOAD_DIR / r["path"]).exists()]  # ìœ íš¨í•œ í•­ëª©ë§Œ
    rows.sort(key=lambda r: r["uploaded_at"], reverse=True)        # ìµœì‹  ë¨¼ì €
    _save_manifest(rows)
    return rows

def save_upload(file) -> dict:
    ts = time.strftime("%Y%m%d_%H%M%S")
    safe_name = f"{ts}__{_slug(file.name)}"
    path = UPLOAD_DIR / safe_name
    with open(path, "wb") as f:
        f.write(file.getbuffer())

    rows = _load_manifest()
    # ë™ì¼ ì›ë³¸ ì´ë¦„ì€ ìµœì‹ ë³¸ìœ¼ë¡œ ì¹˜í™˜(ì´ë¦„ ê¸°ì¤€ ì¤‘ë³µ ì œê±°)
    rows = [r for r in rows if r["name"] != file.name]
    rec = {
        "id": f"{ts}_{int(time.time()*1000)}",
        "name": file.name,
        "path": safe_name,
        "uploaded_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    rows.append(rec)
    _save_manifest(rows)
    return rec

def delete_upload(rec_id: str):
    rows = _load_manifest()
    remain = []
    for r in rows:
        if r["id"] == rec_id:
            try:
                (UPLOAD_DIR / r["path"]).unlink(missing_ok=True)
            except Exception:
                pass
        else:
            remain.append(r)
    _save_manifest(remain)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ ì—…ë¡œë“œ íŒŒì¼ íŒë… â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
@st.cache_data(show_spinner=False)
def read_any(_uploaded_file) -> pd.DataFrame:
    """UploadedFile â†’ DataFrame (ê·¹ë‚´êµ¬ì„± íŒë…ê¸°)"""
    import io, csv
    from bs4 import BeautifulSoup

    raw = _uploaded_file.read()
    if not raw:
        raise RuntimeError("ë¹ˆ íŒŒì¼ì…ë‹ˆë‹¤.")
    name = _uploaded_file.name.lower()
    head = raw[:8]

    def is_zip(h: bytes):   return h.startswith(b"PK\x03\x04")
    def is_ole(h: bytes):   return h.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1")

    # 1) pandas + openpyxl
    if is_zip(head) or name.endswith((".xlsx",".xlsm",".xltx")):
        try:
            df = pd.read_excel(io.BytesIO(raw), engine="openpyxl")
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception:
            pass
        # 2) openpyxl ìˆ˜ë™ íŒŒì‹±
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            header_idx = None
            for i, r in enumerate(rows[:20]):
                non_empty = sum(1 for x in r if (x is not None and str(x).strip() != ""))
                if non_empty >= 2:
                    header_idx = i; break
            if header_idx is None:
                header_idx = 0
            header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
            data = rows[header_idx+1:]
            df = pd.DataFrame(data, columns=header)
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception:
            pass

    # 3) xls / xlsb
    if is_ole(head) or name.endswith(".xls"):
        try:
            df = pd.read_excel(io.BytesIO(raw), engine="xlrd")  # xlrd==1.2.0
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception:
            pass
    if name.endswith(".xlsb"):
        try:
            df = pd.read_excel(io.BytesIO(raw), engine="pyxlsb")
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception:
            pass

    # 4) HTML / MHTML / XML
    try:
        text = raw.decode("utf-8", errors="ignore")
    except Exception:
        text = ""
    if ("<html" in text.lower() or "mime-version" in text.lower()
        or name.endswith((".htm",".html",".xml",".mht",".mhtml"))):
        try:
            tables = pd.read_html(io.StringIO(text))
            if tables:
                df = tables[0]; df.columns = [str(c).strip() for c in df.columns]
                return df
        except Exception:
            try:
                soup = BeautifulSoup(text, "html.parser")
                table = soup.find("table")
                if table:
                    tables = pd.read_html(str(table))
                    if tables:
                        df = tables[0]; df.columns = [str(c).strip() for c in df.columns]
                        return df
            except Exception:
                pass

    # 5) CSV (ì¸ì½”ë”©/êµ¬ë¶„ì/sep= ìë™)
    try:
        head_txt = raw[:256].decode("utf-8-sig", errors="ignore")
        if head_txt.lower().startswith("sep="):
            sep = head_txt.split("=",1)[1].strip()[:1] or ","
            df = pd.read_csv(io.BytesIO(raw), encoding="utf-8-sig",
                             sep=sep, skiprows=1, engine="python", on_bad_lines="skip")
            df.columns = [str(c).strip() for c in df.columns]
            return df
    except Exception:
        pass

    encodings = ["utf-8-sig","cp949","ms949","euc-kr","utf-8"]
    try:
        sample = raw[:4096].decode("utf-8", errors="ignore")
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t")
            delims = [dialect.delimiter]
        except Exception:
            delims = [",",";","\t","|"]
    except Exception:
        delims = [",",";","\t","|"]

    for enc in encodings:
        for delim in delims:
            for skip in range(0, 6):
                try:
                    df = pd.read_csv(io.BytesIO(raw), encoding=enc, sep=delim,
                                     skiprows=skip, engine="python", on_bad_lines="skip")
                    if df.shape[1] >= 2:
                        df.columns = [str(c).strip() for c in df.columns]
                        return df
                except Exception:
                    continue

    sig = head.hex(" ")
    raise RuntimeError(
        "íŒŒì¼ í˜•ì‹ì„ ìë™ìœ¼ë¡œ íŒë…í•˜ì§€ ëª»í–ˆìŠµë‹ˆë‹¤.\n"
        f"- íŒŒì¼ëª…: { _uploaded_file.name }\n"
        f"- í—¤ë” ì‹œê·¸ë‹ˆì²˜: {sig}\n"
        "ì—‘ì…€ì—ì„œ 'ë‹¤ë¥¸ ì´ë¦„ìœ¼ë¡œ ì €ì¥' â†’ .xlsx (í†µí•©ë¬¸ì„œ)ë¡œ ì €ì¥í•˜ê±°ë‚˜, ë”ì¡´ì—ì„œ CSVë¡œ ë‹¤ì‹œ ì¶”ì¶œí•´ ì—…ë¡œë“œ í•´ì£¼ì„¸ìš”.\n"
        "ì•”í˜¸í™”ëœ(ë¹„ë°€ë²ˆí˜¸) í†µí•©ë¬¸ì„œëŠ” ì§€ì›ë˜ì§€ ì•ŠìŠµë‹ˆë‹¤."
    )

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ í‘œì¤€í™” & ë¯¸ì…ê³  ë¡œì§ â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def _pick(df, cands):
    for c in cands:
        if c in df.columns: return c
    return None
def _to_num(s):
    return pd.to_numeric(s, errors="coerce").fillna(0)

def build_base(df: pd.DataFrame) -> pd.DataFrame:
    m = {k: _pick(df, v) for k, v in COL.items()}
    if not m["po_qty"]:
        df["ë°œì£¼ìˆ˜ëŸ‰"] = 0; m["po_qty"] = "ë°œì£¼ìˆ˜ëŸ‰"
    if not m["rcv_qty"]:
        df["ì…ê³ ìˆ˜ëŸ‰"] = 0; m["rcv_qty"] = "ì…ê³ ìˆ˜ëŸ‰"

    base = pd.DataFrame()
    base["ì œí’ˆêµ°"]      = (df[m["group"]].astype(str).str.strip().replace({"": "ë¯¸ì§€ì •"}).fillna("ë¯¸ì§€ì •")
                          if m["group"] else "ë¯¸ì§€ì •")
    base["ë°œì£¼ë²ˆí˜¸"]    = df[m["po_no"]]   if m["po_no"]   else ""
    base["í’ˆëª©ëª…"]      = df[m["item"]]    if m["item"]    else ""
    base["ê±°ë˜ì²˜ëª…"]    = df[m["vendor"]]  if m["vendor"]  else ""
    if m["pgroup"]:
        pg = df[m["pgroup"]].fillna("").astype(str).str.strip()
        base["êµ¬ë§¤ê·¸ë£¹"] = pg
    else:
        base["êµ¬ë§¤ê·¸ë£¹"] = ""
    base["ë°œì£¼ì¼ì"]    = pd.to_datetime(df[m["po_date"]],  errors="coerce") if m["po_date"]  else pd.NaT
    base["ë°œì£¼ë‚©ê¸°ì¼ì"]= pd.to_datetime(df[m["due"]],      errors="coerce") if m["due"]      else pd.NaT
    base["ì…ê³ ì¼ì"]    = pd.to_datetime(df[m["rcv_date"]], errors="coerce") if m["rcv_date"] else pd.NaT
    base["ë°œì£¼ìˆ˜ëŸ‰"]    = _to_num(df[m["po_qty"]])
    base["ì…ê³ ìˆ˜ëŸ‰"]    = _to_num(df[m["rcv_qty"]])
    base["ë¯¸ì…ê³ ìˆ˜ëŸ‰"]  = (base["ë°œì£¼ìˆ˜ëŸ‰"] - base["ì…ê³ ìˆ˜ëŸ‰"]).clip(lower=0)
    base["ì…ê³ êµ¬ë¶„"]    = (df[m["status"]].astype(str).str.strip() if m["status"] else pd.NA)

    stcol = base["ì…ê³ êµ¬ë¶„"].fillna("").astype(str).str.strip()
    base["ìƒíƒœ_í‘œì¤€"] = np.where(stcol.isin(COMPLETE), "ì…ê³ ì™„ë£Œ",
                         np.where(stcol.isin(PARTIAL), "ë¶€ë¶„ì…ê³ ",
                         np.where(stcol.isin(OPEN), "ë¯¸ì…ê³ ",
                                  np.where(stcol=="", "ë¯¸í‘œì‹œ", "ê¸°íƒ€"))))
    return base

def complete_now_series(base: pd.DataFrame) -> pd.Series:
    if base["ì…ê³ êµ¬ë¶„"].notna().any():
        st_series = base["ì…ê³ êµ¬ë¶„"].fillna("").astype(str).str.strip()
        return st_series.isin(COMPLETE)
    return (base["ì…ê³ ìˆ˜ëŸ‰"] >= base["ë°œì£¼ìˆ˜ëŸ‰"])

def complete_by_cutoff(base: pd.DataFrame, cutoff) -> pd.Series:
    cutoff = pd.to_datetime(cutoff)
    return (base["ì…ê³ ìˆ˜ëŸ‰"] >= base["ë°œì£¼ìˆ˜ëŸ‰"]) & base["ì…ê³ ì¼ì"].notna() & (base["ì…ê³ ì¼ì"] <= cutoff)

def backlog_by_cutoff(base: pd.DataFrame, cutoff) -> pd.Series:
    cutoff = pd.to_datetime(cutoff)
    due_ok = base["ë°œì£¼ë‚©ê¸°ì¼ì"].notna() & (base["ë°œì£¼ë‚©ê¸°ì¼ì"] <= cutoff)
    if base["ì…ê³ êµ¬ë¶„"].notna().any():
        st_series = base["ì…ê³ êµ¬ë¶„"].fillna("").astype(str).str.strip()
        return due_ok & st_series.isin(OPEN_OR_PARTIAL)
    else:
        return due_ok & (~complete_by_cutoff(base, cutoff))

def kpi_summary(base: pd.DataFrame, prev_eom: date, curr_eom: date) -> pd.DataFrame:
    total_lines   = base.groupby("ì œí’ˆêµ°").size()
    received_curr = complete_now_series(base).groupby(base["ì œí’ˆêµ°"]).sum()
    prev_overdue  = backlog_by_cutoff(base, prev_eom)
    curr_overdue  = backlog_by_cutoff(base, curr_eom)

    A_cnt = curr_overdue.groupby(base["ì œí’ˆêµ°"]).sum()
    B_cnt = ((~prev_overdue) & curr_overdue).groupby(base["ì œí’ˆêµ°"]).sum()
    prev_cnt = prev_overdue.groupby(base["ì œí’ˆêµ°"]).sum()
    AB_cnt   = A_cnt.add(B_cnt, fill_value=0)

    idx = sorted(total_lines.index.unique().tolist())
    out = pd.DataFrame(index=idx)
    out["ê³„(ì „ì²´)"]         = total_lines
    out["ì…ê³ ê±´ìˆ˜(ë‹¹ì›”ë§)"]  = received_curr.reindex(idx, fill_value=0)
    out["ì „ì›”(ë¯¸ì…ê³ )"]      = prev_cnt.reindex(idx, fill_value=0)
    out["ë‹¹ì›”(A)"]          = A_cnt.reindex(idx, fill_value=0)
    out["ì „ì›” æ¯”"]           = (out["ë‹¹ì›”(A)"] - out["ì „ì›”(ë¯¸ì…ê³ )"]).astype(int)\
                               .map(lambda x: f"â–³{abs(x)}" if x<0 else (f"â–²{x}" if x>0 else "0"))
    out["ì‹ ê·œ(B)"]          = B_cnt.reindex(idx, fill_value=0)
    out["ê³„(A+B)"]          = AB_cnt.reindex(idx, fill_value=0)  # í•„ìš” ì‹œ ì œê±° ê°€ëŠ¥
    out["ì ê¸°ì…ê³ ìœ¨(ë‹¹ì›”ë§)"] = (out["ì…ê³ ê±´ìˆ˜(ë‹¹ì›”ë§)"] / out["ê³„(ì „ì²´)"]).fillna(0.0)

    total = pd.DataFrame(out.sum(numeric_only=True)).T
    total.index = ["í•©ê³„"]
    total["ì „ì›” æ¯”"] = ""
    total["ì ê¸°ì…ê³ ìœ¨(ë‹¹ì›”ë§)"] = out["ì…ê³ ê±´ìˆ˜(ë‹¹ì›”ë§)"].sum() / max(out["ê³„(ì „ì²´)"].sum(), 1)
    return pd.concat([out, total]).reset_index().rename(columns={"index":"ì œí’ˆêµ°"})

def detail_at(base: pd.DataFrame, cutoff: date) -> pd.DataFrame:
    cutoff = pd.to_datetime(cutoff)
    mask = backlog_by_cutoff(base, cutoff)
    D = base.loc[mask].copy()
    if D.empty: return D
    D["ì§€ì—°ì¼ìˆ˜"] = (cutoff - pd.to_datetime(D["ë°œì£¼ë‚©ê¸°ì¼ì"])).dt.days
    cols = ["ì œí’ˆêµ°","ë°œì£¼ë²ˆí˜¸","ê±°ë˜ì²˜ëª…","êµ¬ë§¤ê·¸ë£¹","í’ˆëª©ëª…","ë°œì£¼ì¼ì","ë°œì£¼ë‚©ê¸°ì¼ì",
            "ì…ê³ ì¼ì","ë°œì£¼ìˆ˜ëŸ‰","ì…ê³ ìˆ˜ëŸ‰","ë¯¸ì…ê³ ìˆ˜ëŸ‰","ì…ê³ êµ¬ë¶„","ì§€ì—°ì¼ìˆ˜"]
    return D[cols].sort_values(["ì§€ì—°ì¼ìˆ˜","ë°œì£¼ë‚©ê¸°ì¼ì"], ascending=[False, True])

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ ì—‘ì…€ ë³´ê³ ì„œ ìƒì„± (ì›ë³¸ test2.py í˜•ì‹ ìœ ì§€) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def build_excel(summary, raw_df, prev_eom, curr_eom, det_prev, det_curr) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active; ws.title = "ì›”ë§ìš”ì•½"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(row=1, column=1,
            value=f"<ì›”ë§ ë¯¸ì…ê³  ìš”ì•½>  (ì „ì›”ë§: {pd.to_datetime(prev_eom):%Y-%m-%d}, ë‹¹ì›”ë§: {pd.to_datetime(curr_eom):%Y-%m-%d})"
    ).font = Font(size=13, bold=True)

    r1, r2 = 3, 4
    header = PatternFill("solid", fgColor="F2F2F2")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="999999")

    ws.cell(row=r1, column=1, value="ì œí’ˆêµ°")
    ws.cell(row=r1, column=2, value="ê³„(ì „ì²´)")
    ws.cell(row=r1, column=3, value="ì…ê³ ê±´ìˆ˜(ë‹¹ì›”ë§)")
    ws.merge_cells(start_row=r1, start_column=4, end_row=r1, end_column=6); ws.cell(row=r1, column=4, value="ë¯¸ì…ê³ ")
    ws.merge_cells(start_row=r1, start_column=7, end_row=r1, end_column=8); ws.cell(row=r1, column=7, value="ë¯¸ì…ê³ ")
    ws.cell(row=r1, column=9, value="ì ê¸°ì…ê³ ìœ¨(ë‹¹ì›”ë§)")
    for c in [1,2,3,9]: ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
    for i, lab in enumerate(["ì „ì›”(ë¯¸ì…ê³ )","ë‹¹ì›”(A)","ì „ì›” æ¯”","ì‹ ê·œ(B)","ê³„(A+B)"], start=4):
        ws.cell(row=r2, column=i, value=lab)

    for r in (r1, r2):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True); cell.alignment = center
            cell.fill = header
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ordered = ["ì œí’ˆêµ°","ê³„(ì „ì²´)","ì…ê³ ê±´ìˆ˜(ë‹¹ì›”ë§)","ì „ì›”(ë¯¸ì…ê³ )","ë‹¹ì›”(A)","ì „ì›” æ¯”","ì‹ ê·œ(B)","ê³„(A+B)","ì ê¸°ì…ê³ ìœ¨(ë‹¹ì›”ë§)"]
    for rr in dataframe_to_rows(summary[ordered], index=False, header=False): ws.append(rr)

    max_row = ws.max_row
    for c in range(1,10):
        ws.column_dimensions[get_column_letter(c)].width = 14
        for r in range(r2+1, max_row+1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(r2+1, max_row+1):
        for c in [2,3,4,5,7,8]: ws.cell(row=r, column=c).number_format = '#,##0'
        ws.cell(row=r, column=9).number_format = '0.0%'
    ws.freeze_panes = "A5"

    ws_raw = wb.create_sheet("í˜„í™©(ì›ë³¸)")
    if raw_df is None or raw_df.empty:
        ws_raw.cell(row=1, column=1, value="ì›ë³¸ í‘œê°€ ë¹„ì–´ìˆìŠµë‹ˆë‹¤.")
    else:
        ws_raw.append(list(raw_df.columns))
        for rr in dataframe_to_rows(raw_df, index=False, header=False): ws_raw.append(rr)
        for c in range(1, len(raw_df.columns)+1): ws_raw.cell(row=1, column=c).font = Font(bold=True)
        ws_raw.freeze_panes = "A2"

    ws_p = wb.create_sheet("ë¯¸ì…ê³ _ìƒì„¸(ì „ì›”ë§)")
    if det_prev.empty:
        ws_p.cell(row=1, column=1, value="ì „ì›”ë§ ê¸°ì¤€ ë¯¸ì…ê³ ê°€ ì—†ìŠµë‹ˆë‹¤.")
    else:
        ws_p.append(list(det_prev.columns))
        for rr in dataframe_to_rows(det_prev, index=False, header=False): ws_p.append(rr)
        for c in range(1, len(det_prev.columns)+1): ws_p.cell(row=1, column=c).font = Font(bold=True)
        ws_p.freeze_panes = "A2"

    ws_c = wb.create_sheet("ë¯¸ì…ê³ _ìƒì„¸(ë‹¹ì›”ë§)")
    if det_curr.empty:
        ws_c.cell(row=1, column=1, value="ë‹¹ì›”ë§ ê¸°ì¤€ ë¯¸ì…ê³ ê°€ ì—†ìŠµë‹ˆë‹¤.")
    else:
        ws_c.append(list(det_curr.columns))
        for rr in dataframe_to_rows(det_curr, index=False, header=False): ws_c.append(rr)
        for c in range(1, len(det_curr.columns)+1): ws_c.cell(row=1, column=c).font = Font(bold=True)
        ws_c.freeze_panes = "A2"

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.read()

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ ì‚¬ì´ë“œë°”: ì—…ë¡œë“œ & íˆìŠ¤í† ë¦¬ â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.sidebar.header("ğŸ“¦ ë°ì´í„° ì—…ë¡œë“œ & ê¸°ì¤€ì¼")

# 1) ìƒˆ íŒŒì¼ ì—…ë¡œë“œ(ë‹¤ì¤‘) â†’ ./uploads ì €ì¥ + manifest ê°±ì‹ 
upfiles = st.sidebar.file_uploader(
    "ë°œì£¼í˜„í™© íŒŒì¼ ì—…ë¡œë“œ (.xlsx/.xls/.xlsb/.csv/HTML/XML)",
    type=["xlsx","xls","xlsm","xltx","xlsb","csv","htm","html","xml","mht","mhtml"],
    accept_multiple_files=True
)
if upfiles:
    for f in upfiles:
        rec = save_upload(f)
    st.sidebar.success("ì—…ë¡œë“œ/ì €ì¥ ì™„ë£Œ! (ì¢Œì¸¡ 'ì—…ë¡œë“œ íˆìŠ¤í† ë¦¬'ì— ë°˜ì˜ë¨)")

# 2) ì—…ë¡œë“œ íˆìŠ¤í† ë¦¬: ì„ íƒí•´ì„œ ë¶ˆëŸ¬ì˜¤ê¸° & ì‚­ì œ
st.sidebar.subheader("ğŸ“‚ ì—…ë¡œë“œ íˆìŠ¤í† ë¦¬")
hist = list_uploads()
if hist:
    labels = [f"{i+1}. {h['name']}  Â·  {h['uploaded_at']}" for i, h in enumerate(hist)]
    idx = st.sidebar.selectbox("ì´ì „ ì—…ë¡œë“œ ë¶ˆëŸ¬ì˜¤ê¸°", range(len(hist)), format_func=lambda i: labels[i])
    col_a, col_b = st.sidebar.columns([1,1])
    use_hist   = col_a.button("ì´ íŒŒì¼ ë¶ˆëŸ¬ì˜¤ê¸°")
    del_hist   = col_b.button("ì„ íƒ íŒŒì¼ ì‚­ì œ")
    if del_hist:
        delete_upload(hist[idx]["id"])
        st.sidebar.warning("ì„ íƒ íŒŒì¼ì„ ì‚­ì œí–ˆìŠµë‹ˆë‹¤. (ìƒˆë¡œê³ ì¹¨ ì‹œ ëª©ë¡ ë°˜ì˜)")
else:
    st.sidebar.info("ì €ì¥ëœ ì—…ë¡œë“œê°€ ì—†ìŠµë‹ˆë‹¤.")

# 3) ê¸°ì¤€ì¼
query_date = st.sidebar.date_input("ì¡°íšŒ ê¸°ì¤€ì¼", date.today())
curr_eom = month_end(add_months(query_date, -1))
prev_eom = month_end(add_months(query_date, -2))
st.sidebar.info(f"ì „ì›”ë§: **{prev_eom}**, ë‹¹ì›”ë§: **{curr_eom}**")

# 4) ìƒíƒœ ë¼ë²¨ ì»¤ìŠ¤í„°ë§ˆì´ì¦ˆ
with st.sidebar.expander("ìƒíƒœ ë¼ë²¨ ì»¤ìŠ¤í„°ë§ˆì´ì¦ˆ", expanded=False):
    comp_str = st.text_input("ì…ê³ ì™„ë£Œ ë¼ë²¨(ì‰¼í‘œ)", "ì…ê³ ì™„ë£Œ,ì™„ë£Œ")
    part_str = st.text_input("ë¶€ë¶„ì…ê³  ë¼ë²¨(ì‰¼í‘œ)", "ë¶€ë¶„ì…ê³ ,ë¶€ë¶„,ë¶€ë¶„ì™„ë£Œ")
    open_str = st.text_input("ë¯¸ì…ê³  ë¼ë²¨(ì‰¼í‘œ)", "ë¯¸ì…ê³ ,ëŒ€ê¸°,ë¯¸ì™„ë£Œ")
COMPLETE = set(s.strip() for s in comp_str.split(",") if s.strip())
PARTIAL  = set(s.strip() for s in part_str.split(",") if s.strip())
OPEN     = set(s.strip() for s in open_str.split(",") if s.strip())
OPEN_OR_PARTIAL = OPEN | PARTIAL

st.title("ğŸ“Š ë¯¸ì…ê³  KPI ëŒ€ì‹œë³´ë“œ")

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ ë°ì´í„° ë¡œë”© â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def read_path(path: Path) -> pd.DataFrame:
    class _MemUpload:
        def __init__(self, name, raw): self.name, self._raw = name, raw
        def read(self): return self._raw
    raw = path.read_bytes()
    return read_any(_MemUpload(path.name, raw))

raw_df = None
if hist:
    chosen = hist[idx]
    if use_hist or True:  # ê¸°ë³¸ìœ¼ë¡œ ì„ íƒëœ í•­ëª© ì‚¬ìš©
        try:
            raw_df = read_path(UPLOAD_DIR / chosen["path"])
        except Exception as e:
            st.error(f"íˆìŠ¤í† ë¦¬ íŒŒì¼ ì½ê¸° ì˜¤ë¥˜: {e}")
            halt_app()
else:
    st.info("ì¢Œì¸¡ì—ì„œ íŒŒì¼ì„ ì—…ë¡œë“œí•˜ê±°ë‚˜ íˆìŠ¤í† ë¦¬ì—ì„œ ì„ íƒí•´ ì£¼ì„¸ìš”.")
    halt_app()

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ í‘œì¤€í™”Â·í•„í„°Â·ì§€í‘œ â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
if raw_df is None:
    st.error("ë°ì´í„°ë¥¼ ë¶ˆëŸ¬ì˜¤ì§€ ëª»í–ˆìŠµë‹ˆë‹¤. íŒŒì¼ì„ ì—…ë¡œë“œ í›„ ë‹¤ì‹œ ì‹œë„í•´ ì£¼ì„¸ìš”.")
    halt_app()

base = build_base(raw_df.copy())

st.subheader("ğŸ” í•„í„°")
c1, c2, c3, c4, c5 = st.columns(5)
with c1:
    prods = st.multiselect("ì œí’ˆêµ°", sorted(base["ì œí’ˆêµ°"].dropna().unique().tolist()))
with c2:
    vendors = st.multiselect("ê±°ë˜ì²˜ëª…", sorted(base["ê±°ë˜ì²˜ëª…"].astype(str).dropna().unique().tolist())[:5000])
with c3:
    pg_opts = sorted(
        x for x in base["êµ¬ë§¤ê·¸ë£¹"].fillna("").astype(str).str.strip().unique().tolist()
        if x and x.lower() != "nan"
    )
    purchase_groups = st.multiselect("êµ¬ë§¤ê·¸ë£¹", pg_opts)
with c4:
    statuses = st.multiselect("ì…ê³ êµ¬ë¶„", sorted(base["ì…ê³ êµ¬ë¶„"].astype(str).dropna().unique().tolist()))
with c5:
    state_std = st.multiselect("ìƒíƒœ(í‘œì¤€)", ["ì…ê³ ì™„ë£Œ","ë¶€ë¶„ì…ê³ ","ë¯¸ì…ê³ ","ë¯¸í‘œì‹œ","ê¸°íƒ€"])

due_range = None
order_range = None
due_range_default = None
order_range_default = None
date_cols = st.columns(2)
with date_cols[0]:
    due_series = base["ë°œì£¼ë‚©ê¸°ì¼ì"].dropna()
    if not due_series.empty:
        due_range_default = (due_series.min().date(), due_series.max().date())
        due_range = st.date_input(
            "ë°œì£¼ë‚©ê¸°ì¼ì ë²”ìœ„",
            value=due_range_default,
            min_value=due_range_default[0],
            max_value=due_range_default[1],
            format="YYYY-MM-DD",
        )
    else:
        st.date_input(
            "ë°œì£¼ë‚©ê¸°ì¼ì ë²”ìœ„",
            value=(date.today(), date.today()),
            format="YYYY-MM-DD",
            disabled=True,
            help="ë°œì£¼ë‚©ê¸°ì¼ìê°€ ì—†ì–´ í•„í„°ë¥¼ ì‚¬ìš©í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.",
        )
with date_cols[1]:
    order_series = base["ë°œì£¼ì¼ì"].dropna()
    if not order_series.empty:
        order_range_default = (order_series.min().date(), order_series.max().date())
        order_range = st.date_input(
            "ë°œì£¼ì¼ì ë²”ìœ„",
            value=order_range_default,
            min_value=order_range_default[0],
            max_value=order_range_default[1],
            format="YYYY-MM-DD",
        )
    else:
        st.date_input(
            "ë°œì£¼ì¼ì ë²”ìœ„",
            value=(date.today(), date.today()),
            format="YYYY-MM-DD",
            disabled=True,
            help="ë°œì£¼ì¼ìê°€ ì—†ì–´ í•„í„°ë¥¼ ì‚¬ìš©í•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤.",
        )

flt = base.copy()
if prods:    flt = flt[flt["ì œí’ˆêµ°"].isin(prods)]
if vendors:  flt = flt[flt["ê±°ë˜ì²˜ëª…"].astype(str).isin(vendors)]
if purchase_groups:
    flt = flt[flt["êµ¬ë§¤ê·¸ë£¹"].fillna("").astype(str).str.strip().isin(purchase_groups)]
if statuses: flt = flt[flt["ì…ê³ êµ¬ë¶„"].astype(str).isin(statuses)]
if state_std:flt = flt[flt["ìƒíƒœ_í‘œì¤€"].isin(state_std)]
if due_range_default and isinstance(due_range, (list, tuple)) and len(due_range) == 2:
    due_start, due_end = due_range
    if due_start and due_end:
        due_start_ts = pd.to_datetime(due_start)
        due_end_ts = pd.to_datetime(due_end)
        due_mask = flt["ë°œì£¼ë‚©ê¸°ì¼ì"].notna() & flt["ë°œì£¼ë‚©ê¸°ì¼ì"].between(due_start_ts, due_end_ts)
        if tuple(due_range) == due_range_default:
            flt = flt[due_mask | flt["ë°œì£¼ë‚©ê¸°ì¼ì"].isna()]
        else:
            flt = flt[due_mask]
if order_range_default and isinstance(order_range, (list, tuple)) and len(order_range) == 2:
    order_start, order_end = order_range
    if order_start and order_end:
        order_start_ts = pd.to_datetime(order_start)
        order_end_ts = pd.to_datetime(order_end)
        order_mask = flt["ë°œì£¼ì¼ì"].notna() & flt["ë°œì£¼ì¼ì"].between(order_start_ts, order_end_ts)
        if tuple(order_range) == order_range_default:
            flt = flt[order_mask | flt["ë°œì£¼ì¼ì"].isna()]
        else:
            flt = flt[order_mask]

st.caption(f"í•„í„° ì ìš© ê²°ê³¼: {len(flt):,} í–‰")

summary = kpi_summary(flt, prev_eom, curr_eom)
total_row = summary[summary["ì œí’ˆêµ°"]=="í•©ê³„"].iloc[0]

m = st.columns(6)
m[0].metric("ê³„(ì „ì²´)", f"{int(total_row['ê³„(ì „ì²´)']):,}")
m[1].metric("ì…ê³ ê±´ìˆ˜(ë‹¹ì›”ë§)", f"{int(total_row['ì…ê³ ê±´ìˆ˜(ë‹¹ì›”ë§)']):,}")
m[2].metric("ì „ì›”(ë¯¸ì…ê³ )", f"{int(total_row['ì „ì›”(ë¯¸ì…ê³ )']):,}")
m[3].metric("ë‹¹ì›”(A)", f"{int(total_row['ë‹¹ì›”(A)']):,}")
m[4].metric("ì‹ ê·œ(B)", f"{int(total_row['ì‹ ê·œ(B)']):,}")
m[5].metric("ì ê¸°ì…ê³ ìœ¨", f"{float(total_row['ì ê¸°ì…ê³ ìœ¨(ë‹¹ì›”ë§)']):.1%}")

st.subheader("ğŸ“ˆ ì‹œê°í™”")
curr_over_mask = backlog_by_cutoff(flt, curr_eom)
bar_df = flt.loc[curr_over_mask].groupby("ì œí’ˆêµ°").size().reset_index(name="ë¯¸ì…ê³ (ë‹¹ì›”ë§)")
st.plotly_chart(px.bar(bar_df.sort_values("ë¯¸ì…ê³ (ë‹¹ì›”ë§)", ascending=False),
                       x="ì œí’ˆêµ°", y="ë¯¸ì…ê³ (ë‹¹ì›”ë§)", text_auto=True,
                       title=f"ì œí’ˆêµ°ë³„ ë¯¸ì…ê³ (ë‹¹ì›”ë§: {curr_eom})"),
                use_container_width=True)

topN = st.slider("ê±°ë˜ì²˜ Top N (ë‹¹ì›”ë§ ë¯¸ì…ê³ )", 5, 30, 10)
vendor_df = flt.loc[curr_over_mask].groupby("ê±°ë˜ì²˜ëª…").size().reset_index(name="ë¯¸ì…ê³ (ë‹¹ì›”ë§)")
vendor_df = vendor_df.sort_values("ë¯¸ì…ê³ (ë‹¹ì›”ë§)", ascending=False).head(topN)
st.plotly_chart(px.bar(vendor_df, x="ê±°ë˜ì²˜ëª…", y="ë¯¸ì…ê³ (ë‹¹ì›”ë§)", text_auto=True, title="ê±°ë˜ì²˜ Top ë¯¸ì…ê³ "),
                use_container_width=True)

status_df = flt["ìƒíƒœ_í‘œì¤€"].value_counts().reset_index()
status_df.columns = ["ìƒíƒœ","ê±´ìˆ˜"]
st.plotly_chart(px.pie(status_df, names="ìƒíƒœ", values="ê±´ìˆ˜", hole=0.5, title="ìƒíƒœ ë¶„í¬"),
                use_container_width=True)

st.subheader("ğŸ“‹ ë‹¹ì›”ë§ ë¯¸ì…ê³  ìƒì„¸")
detail_prev = detail_at(flt, prev_eom)
detail_curr = detail_at(flt, curr_eom)
st.dataframe(detail_curr.head(200), use_container_width=True)

st.subheader("ğŸ“‘ ì›”ë§ ìš”ì•½í‘œ")
disp = summary.copy()
disp["ì ê¸°ì…ê³ ìœ¨(ë‹¹ì›”ë§)"] = (disp["ì ê¸°ì…ê³ ìœ¨(ë‹¹ì›”ë§)"]*100).round(1).astype(str) + "%"
st.dataframe(disp, use_container_width=True)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ ë‹¤ìš´ë¡œë“œ â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
st.subheader("â¬‡ï¸ ë‹¤ìš´ë¡œë“œ")
xlsx_bytes = build_excel(summary, raw_df, prev_eom, curr_eom, detail_prev, detail_curr)
st.download_button(
    "ì—‘ì…€ ë³´ê³ ì„œ ë‹¤ìš´ë¡œë“œ (ìš”ì•½+ìƒì„¸+ì›ë³¸ .xlsx)",
    data=xlsx_bytes,
    file_name=f"ë¯¸ì…ê³ _ì›”ë§ë¦¬í¬íŠ¸_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
st.download_button(
    "í•„í„° ê²°ê³¼ CSV ë‹¤ìš´ë¡œë“œ",
    data=flt.to_csv(index=False).encode("utf-8-sig"),
    file_name=f"ë°œì£¼í˜„í™©_í•„í„°ê²°ê³¼_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
    mime="text/csv",
)
st.caption("â€» ì—…ë¡œë“œí•œ íŒŒì¼ì€ ./uploads í´ë”ì— ì €ì¥ë©ë‹ˆë‹¤. íˆìŠ¤í† ë¦¬ì—ì„œ ì„ íƒ/ì‚­ì œ ê°€ëŠ¥í•©ë‹ˆë‹¤.")
